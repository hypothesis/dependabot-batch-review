import sys
from pathlib import Path
from typing import Optional, TextIO, Union

from blessings import Terminal  # type: ignore
from bs4 import BeautifulSoup, PageElement, Tag
from openpyxl import load_workbook  # type: ignore[import-untyped]
from dataclasses import dataclass
from enum import Enum
import re
import os
import subprocess

from .github_client import GitHubClient  # THIS IS THE CRUCIAL IMPORT


class OutputWriter:
    def __init__(self, output_file_path: Optional[Path] = None):
        self._output_file_path = output_file_path
        self._file_handle: Optional[TextIO] = None
        self._t = Terminal()

    def __enter__(self) -> "OutputWriter":
        if self._output_file_path:
            self._output_file_path.parent.mkdir(parents=True, exist_ok=True)
            self._file_handle = open(self._output_file_path, "w", encoding="utf-8")
        return self

    def __exit__(
        self,
        exc_type: Union[type, None],
        exc_val: Union[BaseException, None],
        exc_tb: Union[object, None],
    ) -> None:
        if self._file_handle:
            self._file_handle.close()

    def write(self, content: str, bold: bool = False) -> None:
        if self._file_handle:
            self._file_handle.write(content)
            self._file_handle.write("\n")
        else:
            if bold:
                sys.stdout.write(self._t.bold(content))
            else:
                sys.stdout.write(content)
            sys.stdout.write("\n")

    def write_heading(self, level: int, text: str) -> None:
        if self._file_handle:
            self._file_handle.write(f"{'#' * level} {text}\n\n")
        else:
            self.write(text, bold=True)

    def write_list_item(self, text: str, indent_level: int = 0) -> None:
        prefix = "  " * indent_level + "- "
        if self._file_handle:
            self._file_handle.write(f"{prefix}{{text}}\n")
        else:
            self.write(f"{prefix}{{text}}")

    def write_code_block(self, code: str, lang: str = "") -> None:
        if self._file_handle:
            self._file_handle.write(f"```\n{lang}\n{code}\n```\n")
        else:
            self.write(code)

    def is_interactive(self) -> bool:
        return self._file_handle is None


class CheckStatus(Enum):
    SUCCESS = 1
    FAILED = 2
    MISSING = 3
    PENDING = 4

    @property
    def description(self) -> str:
        return check_status_descriptions[self]


check_status_descriptions = {
    CheckStatus.SUCCESS: "passed",
    CheckStatus.FAILED: "failed",
    CheckStatus.MISSING: "missing",
    CheckStatus.PENDING: "pending",
}


@dataclass
class DependencyUpdate:
    name: str
    from_version: Optional[str]
    to_version: Optional[str]
    notes: str


@dataclass
class DependencyUpdatePR:
    id: str
    package_type: str
    is_group: bool
    group_name: str
    updates: list[DependencyUpdate]
    url: str
    approved: bool
    check_status: CheckStatus
    merge_method: str
    ghsa_id: Optional[str] = None
    advisory_summary: Optional[str] = None
    advisory_url: Optional[str] = None


@dataclass
class DependencyUpdateDetails:
    group_name: str
    is_group: bool
    updates: list[DependencyUpdate]


@dataclass
class RiskAssessment:
    level: str  # "High", "Medium", "Low"
    reasons: list[str]


def analyze_risk(pr: DependencyUpdatePR) -> RiskAssessment:
    reasons = []
    level = "Low"

    for u in pr.updates:
        if u.from_version and u.to_version:
            from_parts = u.from_version.split(".")
            to_parts = u.to_version.split(".")
            if (
                len(from_parts) > 0
                and len(to_parts) > 0
                and from_parts[0] != to_parts[0]
            ):
                level = "High"
                reasons.append(
                    f"Major version bump from {u.from_version} to {u.to_version}"
                )
                break

    for u in pr.updates:
        notes_lower = u.notes.lower()
        if any(
            keyword in notes_lower
            for keyword in ["breaking change", "security", "vulnerability", "cve"]
        ):
            if level != "High":
                level = "High"
            reasons.append(
                "Keywords like 'breaking change' or 'security' found in release notes."
            )
            break

    if pr.check_status == CheckStatus.FAILED:
        if level != "High":
            level = "High"
        reasons.append("CI checks failed.")
    elif pr.check_status in [CheckStatus.PENDING, CheckStatus.MISSING]:
        if level == "Low":
            level = "Medium"
        reasons.append("CI checks are pending or missing.")

    if pr.check_status == CheckStatus.SUCCESS and level == "Low":
        reasons.append("CI checks passed.")

    if not reasons:
        reasons.append("No specific risk factors identified.")

    return RiskAssessment(level=level, reasons=sorted(list(set(reasons))))


def map_risk_to_priority(risk_level: str) -> str:
    if risk_level == "High":
        return "P1"
    elif risk_level == "Medium":
        return "P2"
    else:
        return "P3"


def _extract_ghsa_details(
    soup: BeautifulSoup,
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    ghsa_id = None
    advisory_summary = None
    advisory_url = None

    details_element = soup.find(
        "details", id=lambda x: x and x.startswith("ghsa-details-")
    )
    if isinstance(details_element, Tag):
        ghsa_id_element = details_element.find(
            "a", href=lambda x: x and "github.com/advisories" in x
        )
        if isinstance(ghsa_id_element, Tag):
            ghsa_id = ghsa_id_element.text.strip()
            advisory_url = str(ghsa_id_element["href"])

        summary_element = details_element.find("summary")
        if isinstance(summary_element, Tag):
            advisory_summary = summary_element.text.strip()

    return ghsa_id, advisory_summary, advisory_url


def parse_dependabot_pr(title: str, body: str) -> DependencyUpdateDetails:
    soup = BeautifulSoup(body, "html.parser")
    title_re = r"Bump (\S+) from (\S+) to (\S+)"
    fields_match = re.search(title_re, title, re.IGNORECASE)

    if fields_match:
        dependency, from_version, to_version = fields_match.groups()
        details = [
            d.get_text()
            for d in soup.find_all("details")
            if not d.get_text().strip().startswith("Dependabot commands and options")
        ]
        return DependencyUpdateDetails(
            group_name=dependency,
            is_group=False,
            updates=[
                DependencyUpdate(
                    name=dependency,
                    from_version=from_version,
                    to_version=to_version,
                    notes="\n\n".join(details),
                )
            ],
        )

    group_title_re = r"Bump the (\S+) group"
    group_title_match = re.search(group_title_re, title, re.IGNORECASE)
    if not group_title_match:
        group_title_match = re.search(r"Bump (.*)", title, re.IGNORECASE)

    if not group_title_match:
        raise ValueError("PR title does not match known patterns")
    (group_title,) = group_title_match.groups()

    update_heading_pat = r"Updates (\S+) from (\S+) to (\S+)"

    def contains_update_heading(el: PageElement) -> bool:
        return re.search(update_heading_pat, el.get_text()) is not None

    headings = [p for p in soup.find_all("p") if contains_update_heading(p)]

    single_update_pat = r"Bumps the \S+ group with 1 update: (\S+)"
    if not headings:
        headings = [
            p for p in soup.find_all("p") if re.match(single_update_pat, p.get_text())
        ]
        if not headings:
            raise ValueError("Package names not found in PR body")

    updates = []
    for heading in headings:
        fields_match = re.search(update_heading_pat, heading.get_text(), re.IGNORECASE)
        if fields_match:
            dependency, from_version, to_version = fields_match.groups()
        else:
            fields_match = re.search(
                single_update_pat, heading.get_text(), re.IGNORECASE
            )
            assert fields_match
            (dependency,) = fields_match.groups()
            from_version = None
            to_version = None

        notes: list[str] = []
        curr = heading.next_sibling
        while curr and not contains_update_heading(curr) and curr.name != "hr":
            if curr.name == "details":
                notes.append(curr.get_text())
            curr = curr.next_sibling

        updates.append(
            DependencyUpdate(
                name=dependency,
                from_version=from_version,
                to_version=to_version,
                notes="\n\n".join(notes),
            )
        )

    return DependencyUpdateDetails(
        group_name=group_title, is_group=True, updates=updates
    )


def parse_package_type_from_branch_name(branch: str) -> str:
    branch_name_re = r"^dependabot/([^/]+)/.*"
    branch_name_match = re.search(branch_name_re, branch)
    if not branch_name_match:
        raise ValueError(f"Failed to parse branch name '{branch}'")
    package_type = branch_name_match.groups()[0]
    return package_type


def fetch_dependency_prs(
    gh: GitHubClient,
    organization: str,
    repo_filter: Optional[str] = None,
    labels: list[str] = ["dependencies"],
) -> list[DependencyUpdatePR]:
    dependencies_query = """
    query($query: String!) {
      search(type:ISSUE, query: $query, first:100) {
        issueCount
        nodes {
          ... on PullRequest {
            repository {
              name
              viewerDefaultMergeMethod
            }

            author { login }
            id
            title
            bodyHTML
            headRefName
            reviewDecision
            url

            commits (last:1) {
              nodes {
                commit {
                  statusCheckRollup {
                    state
                  }
                }
              }
            }
          }
        }
      }
    }
    """

    label_terms = " ".join(f"label:{label}" for label in labels)
    query = f"org:{organization} {label_terms} is:pr is:open author:app/dependabot"
    result = gh.query(query=dependencies_query, variables={"query": query})
    pull_requests = result["search"]["nodes"]

    updates: list[DependencyUpdatePR] = []
    for pr in pull_requests:
        repo = pr["repository"]["name"]

        if repo_filter is not None:
            if repo_filter not in repo:
                continue

        try:
            update_details = parse_dependabot_pr(pr["title"], pr["bodyHTML"])
            status_check_rollup = pr["commits"]["nodes"][0]["commit"][
                "statusCheckRollup"
            ]
            package_type = parse_package_type_from_branch_name(pr["headRefName"])

            soup = BeautifulSoup(pr["bodyHTML"], "html.parser")
            ghsa_id, advisory_summary, advisory_url = _extract_ghsa_details(
                soup
            )  # This will now return None, None, None

        except ValueError as exc:
            print(f"Failed to parse details from {pr['url']}: {exc}", file=sys.stderr)
            continue

        rollup_state = status_check_rollup["state"] if status_check_rollup else None
        if rollup_state == "SUCCESS":
            check_status = CheckStatus.SUCCESS
        elif rollup_state == "PENDING" or rollup_state == "EXPECTED":
            check_status = CheckStatus.PENDING
        elif rollup_state == "ERROR" or rollup_state == "FAILURE":
            check_status = CheckStatus.FAILED
        elif rollup_state is not None:
            check_status = CheckStatus.FAILED
        else:
            check_status = CheckStatus.MISSING

        updates.append(
            DependencyUpdatePR(
                id=pr["id"],
                is_group=update_details.is_group,
                group_name=update_details.group_name,
                approved=pr["reviewDecision"] == "APPROVED",
                check_status=check_status,
                updates=update_details.updates,
                merge_method=pr["repository"]["viewerDefaultMergeMethod"],
                package_type=package_type,
                url=pr["url"],
                ghsa_id=ghsa_id,
                advisory_summary=advisory_summary,
                advisory_url=advisory_url,
            )
        )

    return updates


def merge_pr(gh: GitHubClient, pr_id: str, merge_method: str = "MERGE") -> None:
    merge_query = """
    mutation mergePullRequest($input: MergePullRequestInput!) {
      mergePullRequest(input: $input) {
        pullRequest {
          merged
          url
        }
      }
    }
    """
    gh.query(
        merge_query, {"input": {"pullRequestId": pr_id, "mergeMethod": merge_method}}
    )


class PromptAbortError(Exception):
    pass


def read_action(prompt: str, actions: list[str], default: Optional[str] = None) -> str:
    if not os.isatty(sys.stdout.fileno()) and default:
        return default

    while True:
        try:
            user_input = input(f"{prompt} > ").strip().lower()
        except EOFError as e:
            raise PromptAbortError() from e
        except KeyboardInterrupt as e:
            raise PromptAbortError() from e

        for action in actions:
            if action == user_input:
                return action

        for action in actions:
            if action.startswith(user_input):
                return action


def open_url(url: str) -> None:
    subprocess.call(["open", url])


def get_package_diff(package_type: str, update: DependencyUpdate) -> str | None:
    if not update.from_version or not update.to_version:
        return None

    match package_type:
        case "npm_and_yarn":
            try:
                cmd = [
                    "npm",
                    "diff",
                    "--diff",
                    f"{update.name}@{update.from_version}",
                    "--diff",
                    f"{update.name}@{update.to_version}",
                ]
                print(f"Running command: {{{' '.join(cmd)}}}")
                sys.stdout.flush()
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
                if result.returncode == 0:
                    return result.stdout
                else:
                    return f"Error running npm diff: {{{result.stderr}}}"
            except subprocess.TimeoutExpired:
                return "Error: npm diff command timed out"
            except FileNotFoundError:
                return "Error: npm command not found. Please ensure npm is installed and in your PATH."
            except Exception as e:
                return f"Error running npm diff: {{{str(e)}}}"
        case _:
            return None


def generate_xlsx_report(
    prs: list[DependencyUpdatePR], template_path: Path, output_path: Path
) -> None:
    workbook = load_workbook(template_path)
    if "Alerts" not in workbook.sheetnames:
        raise ValueError(
            f"Template '{template_path}' does not contain an 'Alerts' sheet."
        )

    sheet = workbook["Alerts"]
    header = [cell.value for cell in sheet[1]]
    col_map = {name: idx for idx, name in enumerate(header)}
    next_row = sheet.max_row + 1

    for pr in prs:
        url_parts = pr.url.split("/")
        repo_owner = url_parts[-4]
        repo_name = url_parts[-3]
        full_repo_name = f"{repo_owner}/{repo_name}"
        pr_number = url_parts[-1]

        risk = analyze_risk(pr)
        priority = map_risk_to_priority(risk.level)

        row_values: list[Union[str, int, None]] = [None] * len(header)

        row_values[col_map["Repo"]] = full_repo_name
        row_values[col_map["GHSA"]] = pr.ghsa_id
        row_values[col_map["Alerts Count"]] = len(pr.updates)
        row_values[col_map["Severity"]] = risk.level
        row_values[col_map["Priority"]] = priority
        row_values[col_map["Package"]] = pr.group_name
        row_values[col_map["Ecosystem"]] = pr.package_type
        row_values[col_map["Advisory Summary"]] = pr.advisory_summary
        row_values[col_map["Advisory URL"]] = pr.advisory_url
        row_values[col_map["PR/MR URL"]] = pr.url
        row_values[col_map["PR/MR #"]] = int(pr_number)
        row_values[col_map["Merge Decision"]] = "To Review"
        row_values[col_map["Merged?"]] = "No"
        row_values[col_map["Merged Date"]] = None
        row_values[col_map["Merged By"]] = None
        row_values[col_map["Notes"]] = "\n".join(risk.reasons)

        for col_idx, value in enumerate(row_values):
            sheet.cell(row=next_row, column=col_idx + 1, value=value)

        next_row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def review_updates(
    gh_client: GitHubClient,
    prs: list[DependencyUpdatePR],
    output_md_path: Optional[str] = None,
) -> None:
    output_writer = (
        OutputWriter(Path(output_md_path)) if output_md_path else OutputWriter()
    )

    with output_writer:
        if not output_writer.is_interactive():
            prs_by_group_name: dict[str, list[DependencyUpdatePR]] = {}
            for pr in prs:
                if pr.group_name not in prs_by_group_name:
                    prs_by_group_name[pr.group_name] = []
                prs_by_group_name[pr.group_name].append(pr)

            for group_name, group_prs in prs_by_group_name.items():
                output_writer.write_heading(1, f"Dependency: {{{group_name}}}")

                prs_by_version: dict[str, list[DependencyUpdatePR]] = {}
                for pr in group_prs:
                    version_key = " -> ".join(
                        sorted(
                            list(
                                set(
                                    f"{u.from_version or '?'} -> {u.to_version or '?'}"
                                    for u in pr.updates
                                )
                            )
                        )
                    )
                    if version_key not in prs_by_version:
                        prs_by_version[version_key] = []
                    prs_by_version[version_key].append(pr)

                for version_key, version_prs in prs_by_version.items():
                    output_writer.write_heading(2, f"Version: {{{version_key}}}")

                    for pr in version_prs:
                        output_writer.write_heading(3, f"PR: {{{pr.url}}}")

                        risk = analyze_risk(pr)
                        output_writer.write_heading(4, "Risk Analysis")
                        output_writer.write(f"**Level:** {{{risk.level}}}")
                        output_writer.write("**Reasons:**")
                        for reason in risk.reasons:
                            output_writer.write_list_item(reason)
                        output_writer.write("")

                        output_writer.write(
                            f"**CI Status:** {{{pr.check_status.description}}}"
                        )
                        output_writer.write("")

                        for u in pr.updates:
                            if u.notes:
                                output_writer.write_heading(
                                    4, f"Release Notes for {{{u.name}}}"
                                )
                                output_writer.write_code_block(u.notes)

                        for u in pr.updates:
                            if diff_output := get_package_diff(pr.package_type, u):
                                output_writer.write_heading(4, f"Diff for {{{u.name}}}")
                                output_writer.write_code_block(diff_output, lang="diff")

            return

        version_bumps = set()
        for pr in prs:
            for u in pr.updates:
                version_bumps.add((u.name, u.from_version, u.to_version))

        output_writer.write_heading(2, "Versions")
        for name, from_ver, to_ver in version_bumps:
            from_ver = from_ver or "(unknown)"
            to_ver = to_ver or "(unknown)"
            output_writer.write_list_item(f"{name} {from_ver} -> {to_ver}")

        updates_by_status: dict[CheckStatus, list[DependencyUpdatePR]] = {}
        for update in prs:
            if update.check_status not in updates_by_status:
                updates_by_status[update.check_status] = []
            updates_by_status[update.check_status].append(update)

        check_statuses: list[str] = []
        for status, items in updates_by_status.items():
            check_statuses.append(f"{len(items)} {status.description}")
        output_writer.write(f"Checks: {', '.join(check_statuses)}")

        for update in prs:
            if update.check_status == CheckStatus.SUCCESS:
                continue
            output_writer.write_list_item(
                f"{update.url} checks {update.check_status.description}", indent_level=1
            )

        while True:
            action = read_action(
                prompt="[m]erge passing, [s]kip, [q]uit, [r]eview changes, package [d]iff, [v]iew in browser, [l]ist URLs",
                actions=["diff", "merge", "skip", "quit", "review", "list", "view"],
                default="skip",
            )
            if action == "quit":
                return
            elif action == "merge":
                for update in prs:
                    if update.check_status != CheckStatus.SUCCESS:
                        continue

                    output_writer.write(f"Merging {update.url} …")
                    try:
                        merge_pr(
                            gh_client, pr_id=update.id, merge_method=update.merge_method
                        )
                    except Exception as e:
                        output_writer.write(f"Merge failed: {repr(e)}")
                break
            elif action == "skip":
                break
            elif action == "review":
                notes = []
                for u in prs[0].updates:
                    notes += u.notes.splitlines()
                max_lines = 35
                if len(notes) > max_lines:
                    notes = notes[0:max_lines]
                    notes.append('... (Enter "view" to see full notes in browser)')
                for line in notes:
                    output_writer.write(f"  {line}")
            elif action == "view":
                open_url(prs[0].url)
            elif action == "list":
                urls = sorted(u.url for u in prs)
                for url in urls:
                    output_writer.write(f"  {url}")
            elif action == "diff":
                unique_updates: dict[
                    tuple[str, str | None, str | None], tuple[str, DependencyUpdate]
                ] = {}
                for pr in prs:
                    for pr_update in pr.updates:
                        key = (
                            pr_update.name,
                            pr_update.from_version,
                            pr_update.to_version,
                        )
                        if key not in unique_updates:
                            unique_updates[key] = (pr.package_type, pr_update)

                diffs = []
                for (name, from_version, to_version), (
                    package_type,
                    dep_update,
                ) in unique_updates.items():
                    if diff_output := get_package_diff(package_type, dep_update):
                        diffs.append((name, from_version, to_version, diff_output))

                if not diffs:
                    output_writer.write(
                        """
Package diffs are not available for these packages.

Package diffs are currently only available for npm packages."""
                    )
                else:
                    combined_diff = ""
                    for package_name, from_version, to_version, diff_output in diffs:
                        from_ver = from_version or "(unknown)"
                        to_ver = to_version or "(unknown)"
                        combined_diff += f"\n--- Diff for {package_name} {from_ver} -> {to_ver} ---\n"
                        combined_diff += diff_output + "\n"

                    try:
                        process = subprocess.Popen(
                            ["less"], stdin=subprocess.PIPE, text=True
                        )
                        process.communicate(input=combined_diff)
                    except FileNotFoundError:
                        output_writer.write(
                            "Error: 'less' command not found. Please install less to view diffs."
                        )
                    except subprocess.CalledProcessError:
                        output_writer.write("Error: Failed to display diff with less.")
